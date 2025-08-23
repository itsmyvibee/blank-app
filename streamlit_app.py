import os
from pathlib import Path
import streamlit as st
from openpyxl import load_workbook  # só para descobrir o nome da planilha
from xlcalculator import ModelCompiler, Evaluator

# st.set_page_config(page_title="Excel como motor", page_icon="🧮")

# # Oculta menus/ícone do GitHub (visual)
# st.markdown("""
# <style>
# #MainMenu, footer {visibility: hidden;}
# header [data-testid="stToolbar"] {visibility: hidden;}
# a[href*="github.com"] {display:none !important;}
# </style>
# """, unsafe_allow_html=True)

# st.title("Excel como motor de cálculo (B1, B2 → B3)")

# # === AJUSTE AQUI SE PRECISAR ===
# # Se o arquivo estiver na mesma pasta do streamlit_app.py:
# EXCEL_PATH = Path(__file__).parent / "Book_1.xlsx"
# # Exemplo se estiver em subpasta "assets":
# # EXCEL_PATH = Path(__file__).parent / "assets" / "Book_1.xlsx"
# # ================================

# # Debug: listar arquivos da pasta do app
# st.caption("Arquivos na pasta do app:")
# st.code("\n".join(sorted(os.listdir(Path(__file__).parent))), language="text")

# # Checar existência do arquivo local (sem URL!)
# if not EXCEL_PATH.exists():
#     st.error(f"Arquivo não encontrado: {EXCEL_PATH.name}. "
#              f"Coloque-o no mesmo diretório do app ou ajuste EXCEL_PATH.")
#     st.stop()

# # Descobrir o nome da primeira planilha (caso você não queira fixar)
# try:
#     wb = load_workbook(EXCEL_PATH, data_only=False, read_only=True)
#     SHEET = wb.sheetnames[0]
# except Exception as e:
#     st.error(f"Não consegui abrir o Excel para ler o nome da planilha: {e}")
#     st.stop()

# col1, col2 = st.columns(2)
# with col1:
#     x = st.number_input("Valor para B1", value=0.0, step=1.0, format="%.2f")
# with col2:
#     y = st.number_input("Valor para B2", value=0.0, step=1.0, format="%.2f")

# if st.button("Calcular"):
#     try:
#         # Compilar e avaliar SEM cache, para evitar vestígios de URL antiga
#         model = ModelCompiler().read_and_parse_archive(str(EXCEL_PATH))
#         ev = Evaluator(model)

#         ev.set_cell_value(f"{SHEET}!B1", x)
#         ev.set_cell_value(f"{SHEET}!B2", y)
#         result = ev.evaluate(f"{SHEET}!B3")

#         if isinstance(result, (int, float)):
#             st.success(f"Resultado em B3: {result:.2f}")
#         else:
#             st.warning(f"B3 retornou um valor não numérico: {result}")

#     except Exception as e:
#         st.error(f"Erro ao avaliar B3: {e}")

# INTERFACE TESTE ----------------------------------------------------------------------------------

# INTERFACE TESTE ----------------------------------------------------------------------------------

import streamlit as st

st.set_page_config(page_title="Simulador de P&L", layout="wide")

# ----------------------- ESTILO OPCIONAL -----------------------
st.markdown("""
<style>
.app-title { font-size: 28px; font-weight: 700; letter-spacing: .2px; }
.app-subtitle { color:#888; font-size:18px; margin-bottom: 1.25rem; }
hr.thin { border: none; border-top: 1px solid #eaeaea; margin: 6px 0 18px 0; }
.flag-col { display:flex; align-items:center; height: 42px; font-weight: 600; }
.header-cell { font-weight:600; padding-bottom:6px; border-bottom:2px solid #222; }
.row-sep { border-bottom: 1px solid #e5e5e5; margin: 0 0 8px 0; }
</style>
""", unsafe_allow_html=True)

def brl(v: float) -> str:
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# ----------------------- CABEÇALHO -----------------------
left, _ = st.columns([1,3])
with left:
    st.markdown('<div class="app-title">fiserv.</div>', unsafe_allow_html=True)
    st.markdown('<div class="app-subtitle">Simulador de P&L</div>', unsafe_allow_html=True)

# ============================================================
#  CAMPOS SUPERIORES (FORA DO FORM) -> atualizam em tempo real
# ============================================================
c1, c2, c3 = st.columns([1.2, 1, 1])

with c1:
    nome = st.text_input("Nome do estabelecimento", placeholder="Jane Smith", key="nome")
    cnpj_principal = st.text_input("CNPJ Principal", placeholder="00.000.000/0000-00", key="cnpj_principal")

with c2:
    faturamento_anual = st.number_input("Faturamento Anual (R$)", min_value=0.0, step=1000.0, format="%.2f", key="fat_anual")
    # Mensal calculado e exibido (só leitura)
    fat_mensal_num = st.session_state.get("fat_anual", 0.0) / 12.0
    st.text_input("Faturamento Mensal (R$)", value=brl(fat_mensal_num), disabled=True)

with c3:
    antecipacao_sel = st.selectbox("Antecipação?", ["SIM", "NÃO"], key="antecipacao_sel")
    # >>> Captura FORA do form: muda a UI imediatamente
    captura_sel = st.selectbox("Captura", ["FISICO", "ECOMMERCE"], key="captura_sel")

st.markdown("<hr class='thin'/>", unsafe_allow_html=True)

# ==========================================
#  TABELAS DENTRO DO FORM + SUBMIT
# ==========================================
with st.form("form_pl"):
    # ----------------------- TABELA DE TAXAS -----------------------
    st.markdown("#### Taxas solicitadas")
    h1, h2, h3, h4, h5 = st.columns([1.1, 1, 1, 1, 1])
    with h1: st.markdown('<div class="header-cell"> </div>', unsafe_allow_html=True)
    with h2: st.markdown('<div class="header-cell">Débito</div>', unsafe_allow_html=True)
    with h3: st.markdown('<div class="header-cell">Crédito</div>', unsafe_allow_html=True)
    with h4: st.markdown('<div class="header-cell">Parcelado 2 a 6</div>', unsafe_allow_html=True)
    with h5: st.markdown('<div class="header-cell">Parcelado 7 a 12</div>', unsafe_allow_html=True)

    bandeiras = [
        ("Mastercard", "mc"),
        ("Visa", "visa"),
        ("Elo", "elo"),
        ("American Express", "amex"),
    ]

    taxas = {}
    for nome_bandeira, key_base in bandeiras:
        cA, cB, cC, cD, cE = st.columns([1.1, 1, 1, 1, 1])
        with cA:
            st.markdown(f"<div class='flag-col'>{nome_bandeira}</div>", unsafe_allow_html=True)
        with cB:
            taxas[f"{key_base}_debito"] = st.number_input(
                f"Débito — {nome_bandeira}", min_value=0.0, max_value=5.0, value=0.0, step=0.10,
                format="%.2f", key=f"{key_base}_deb", label_visibility="collapsed"
            )
        with cC:
            taxas[f"{key_base}_credito"] = st.number_input(
                f"Crédito — {nome_bandeira}", min_value=0.0, max_value=5.0, value=0.0, step=0.10,
                format="%.2f", key=f"{key_base}_cred", label_visibility="collapsed"
            )
        with cD:
            taxas[f"{key_base}_parc_2a6"] = st.number_input(
                f"Parcelado 2 a 6 — {nome_bandeira}", min_value=0.0, max_value=5.0, value=0.0, step=0.10,
                format="%.2f", key=f"{key_base}_p26", label_visibility="collapsed"
            )
        with cE:
            taxas[f"{key_base}_parc_7a12"] = st.number_input(
                f"Parcelado 7 a 12 — {nome_bandeira}", min_value=0.0, max_value=5.0, value=0.0, step=0.10,
                format="%.2f", key=f"{key_base}_p712", label_visibility="collapsed"
            )
        #st.markdown("<div class='row-sep'></div>", unsafe_allow_html=True)

    # ----------------------- TABELA DE TERMINAIS (CONDICIONAL) -----------------------
    terminais_data = {}
    if st.session_state.get("captura_sel") == "FISICO":   # aparece/desaparece na hora
        st.markdown("#### ")
        st.markdown("<div class='row-sep'></div>", unsafe_allow_html=True)
        st.markdown("#### Tecnlogia Física")
        t1, t2, t3, _sp2 = st.columns([1.4, 1, 1, 3])
        with t1: st.markdown('<div class="header-cell"></div>', unsafe_allow_html=True)
        with t2: st.markdown('<div class="header-cell">Quantidade</div>', unsafe_allow_html=True)
        with t3: st.markdown('<div class="header-cell">Valor (R$)</div>', unsafe_allow_html=True)

        lista_terminais = [
            ("POS Wifi", "pos_wifi"),
            ("PINPAD", "pinpad"),
            ("SMART POS", "smart_pos"),
            ("CLOVER FLEX", "clover_flex"),
            ("CLOVER MINI", "clover_mini"),
        ]
        for label, slug in lista_terminais:
            cA, cB, cC, spacer2 = st.columns([1.4, 1, 1, 3])
            with cA:
                st.write(label)
            with cB:
                q = st.number_input(
                    f"Qtd — {label}", min_value=0, value=0, step=1,
                    key=f"{slug}_qtd", label_visibility="collapsed"
                )
            with cC:
                v = st.number_input(
                    f"Valor — {label}", min_value=0.0, value=0.0, step=10.0, format="%.2f",
                    key=f"{slug}_valor", label_visibility="collapsed"
                )
            terminais_data[slug] = {"terminal": label, "quantidade": q, "valor": v}
            #st.markdown("<div class='row-sep'></div>", unsafe_allow_html=True)

    # ----------------------- Campo de antecipação (CONDICIONAL) -----------------------
    if st.session_state.get("antecipacao_sel") == "SIM":   # aparece/desaparece na hora
        st.markdown("#### ")
        st.markdown("<div class='row-sep'></div>", unsafe_allow_html=True)
        st.markdown("#### Antecipação")
        taxa_antecipacao = st.number_input("Taxa de antecipação (%)", min_value=0.0, step=0.10, format="%.2f", key="taxa_antecipacao", label_visibility="collapsed")

    
    # ----------------------- SUBMIT -----------------------
    submitted = st.form_submit_button("Submit")

# ----------------------- PÓS-SUBMIT: MOSTRAR DADOS -----------------------
if submitted:
    resultado = {
        "estabelecimento": st.session_state.get("nome"),
        "cnpj_principal": st.session_state.get("cnpj_principal"),
        "faturamento_anual": st.session_state.get("fat_anual", 0.0),
        "faturamento_mensal": fat_mensal_num,
        "antecipacao": st.session_state.get("antecipacao_sel"),
        "captura": st.session_state.get("captura_sel"),
        "taxas_por_bandeira_percent": taxas,
        "terminais": terminais_data,  # vazio se não for FISICO
        "prepay_tax":st.session_state.get("taxa_antecipacao", 0.0)
    }
    st.success("Dados coletados com sucesso!")
    st.json(resultado)


